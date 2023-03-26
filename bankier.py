import os
import requests
from bs4 import BeautifulSoup
from tickers import tickers_list
from openpyxl import load_workbook

lista = tickers_list()
location = os.getenv("TICKERS_PATH")

workbook = load_workbook(filename=location)
sheet_1 = workbook['Freefloat']

def run_example():

    LICZBA_SPACJI = 73
    MAX_TRIES = 3
    row_number = 1

    for ticker in lista:
        url = "https://www.bankier.pl/gielda/notowania/new-connect/" + ticker + "/akcjonariat"
        print(url)
        print(ticker)
        tries = 0
        while tries < MAX_TRIES:

            try:
                response = requests.get(url)
                response.raise_for_status()
                break  # breaks out of the while loop if successful
            except requests.RequestException as error:
                tries += 1
                print(f"Connection error: {error}. Trying again ({tries}/{MAX_TRIES})...")

        else:
            # If all tries failed
            print(f"Request failed after {MAX_TRIES} attempts. Skipping {ticker}.")
            continue

        with open(f"{ticker}.html", "w", encoding="utf-8") as file:
            file.write(response.text)

        text = response.text
        soup = BeautifulSoup(text, "html.parser")

        # Find the "Free float:" in the table
        free_float_row = soup.find("td", text="Free float:").parent

        # Extract the value from the fourth column
        free_float_value = free_float_row.find_all("td")[3].get_text(strip=True)

        print("FREE FLOAT VALUE..................")
        print(free_float_value)
        input("cos")

        # sheet_1.cell(row=row_number, column=1, value=ticker)
        # sheet_1.cell(row=row_number, column=2, value=wanted_text)
        # sheet_1.cell(row=row_number, column=3, value=url)
        #
        # row_number += 1
    workbook.save(location)


if __name__ == "__main__":
    run_example()
