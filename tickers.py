"""A script that return list of tickers of WSE listed companies"""
import os
from openpyxl import load_workbook

tickers_location = os.getenv("TICKERS")
print(tickers_location)

workbook = load_workbook(filename=tickers_location)
first_sheet = workbook.active

rows = first_sheet.max_row + 1


def get_tickers_list():
    """Returns list of tickers"""
    tickers_list = []
    for i in range(1, rows):
        tickers_list.append(first_sheet[("A" + str(i))].value)
    return tickers_list
